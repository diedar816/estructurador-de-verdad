# -*- coding: utf-8 -*-
"""
Estructurador de Información (PDF -> Excel) – FIX3+7
-------------------------------------------------------------------------------
Correcciones sobre FIX3+6 sin perder mejoras previas (probado contra el PDF
"Estructurador de Información3.pdf") y corrigiendo los errores observados en
"Estructurador de Información1.pdf":

1) **Lugar de los hechos**
   - Si el detector lo marcaba como DUMMY (p.ej., cad. con muchos MAYÚS./slashes
     como "11001 BARRIO/LOCALIDAD/COMUNA:..." o un simple código DANE "05001"),
     antes quedaba **vacío**. Ahora:
       * Se aplica un **fallback inteligente** que intenta sintetizar el lugar
         en formato legible ("CIUDAD BERNA, ANTONIO NARIÑO, BOGOTÁ, D.C.").
       * Si solo hay **código DANE** (p. ej. 11001/05001), se deja "<código>
         (<Municipio>)" usando un pequeño catálogo DANE integrado.
       * Se mantiene la bandera de calidad (`Quality_Lugar de los hechos`)
         como "HEURISTIC" o "CODE_ONLY".

2) **Teléfono de notificación**
   - En algunos PDF venía vacío y el extractor de vecindad capturaba el número
     del **Teléfono móvil** (siguiente etiqueta). Ahora la búsqueda de vecindad
     está **acotada entre la etiqueta actual y la siguiente etiqueta**; si aun
     así ambos quedan iguales, el **Teléfono de notificación se deja vacío**.

3) **CLI opcional**
   - Permite `-i/--input` (PDF) y `-o/--output` (XLSX). Defaults conservadores.

Resto de funcionalidades de FIX3+6 permanecen intactas: normalizador de tipo de
Documento, N° de documento a dígitos, orden por rol (INDICIADO primero), etc.
"""
import os, re, sys, time, traceback, datetime, unicodedata, argparse
from pathlib import Path
import pandas as pd
from PyPDF2 import PdfReader

# ---------------- Excel ----------------
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

# ---------------- Configuración ----------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_PDF_NAME = "Estructurador de Información.pdf"
DEFAULT_XLSX_NAME = "Estructurado en tabla.xlsx"
DEDUPE_BY_ID = False

# --- OCR opcional ---
TESSERACT_EXE = r"C:\\Program Files\\Tesseract-OCR\\tesseract.exe"
POPPLER_PATH = None
OCR_OK = True
try:
    import pytesseract
    from pdf2image import convert_from_path
    from PIL import Image, ImageOps, ImageFilter
    if TESSERACT_EXE and os.path.exists(TESSERACT_EXE):
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_EXE
except Exception:
    OCR_OK = False

# ---------------- Definición de campos ----------------
CASE_FIELDS = [
    "Caso Noticia","Ley de Aplicabilidad","Procedimiento Abreviado?","Priorizado",
    "Tipo Noticia","Delito","Grado Delito","Caracterización","Modalidad","Modo",
    "Fecha de los Hechos","Lugar de los hechos","Relato de los hechos",
    "Municipio Fiscal","Seccional","Unidad de Fiscalía","Despacho",
    "Estado de la asignación","Unidad de Enrutamiento","Estado del caso","Etapa del caso",
]

PERSON_FIELDS = [
    "Calidad","Documento","Número documento","Nombre",
    "Departamento de notificación","Municipio de notificación","Dirección de notificación",
    "Teléfono de notificación","Teléfono móvil","Correo Electrónico","Teléfono Oficina",
]

LABEL_ALIASES = {
    "Ley de Aplicabilidad": [r"Ley de\s+Aplicabilidad"],
    "Procedimiento Abreviado?": [r"Procedimiento\s+Abreviado\s*\??"],
    "Relato de los hechos": [r"Relato de los\s+hechos"],
    "Fecha de los Hechos": [r"Fecha de los\s+Hechos"],
    "Lugar de los hechos": [r"Lugar de los\s+hechos"],
    "Unidad de Fiscalía": [r"Unidad de\s+Fiscal[ií]a"],
    "Correo Electrónico": [r"Correo\s+Electr[oó]nico"],
    # Número documento — variantes
    "Número documento": [
        r"N[uú]mero\s+documento",
        r"N[uú]mero\s+de\s+documento",
        r"Documento\s*No\.",
        r"No\.?\s*de\s*documento",
        r"Identificaci[oó]n\s*No\.?"
    ],
    "Estado de la asignación": [r"Estado de la\s+asignaci[oó]n"],
    "Unidad de Enrutamiento": [r"Unidad de\s+Enrutamiento"],
    "Teléfono móvil": [r"Tel[eé]fono\s+m[oó]vil", r"Celular"],
    "Teléfono de notificación": [r"Tel[eé]fono\s+de\s+notificaci[oó]n"],
    "Dirección de notificación": [r"Direcci[oó]n\s+de\s+notificaci[oó]n"],
    "Seccional": [r"Direcci[oó]n\s+Seccional"],
    "Departamento de notificación": [r"Departamento\s+de\s+notificaci[oó]n"],
    "Municipio de notificación": [r"Municipio\s+de\s+notificaci[oó]n"],
    # Sinónimos para Documento (tipo)
    "Documento": [
        r"Tipo\s+de\s+documento",
        r"Tipo\s*Documento",
        r"Tipo\s+de\s+identificaci[oó]n",
        r"Tipo\s+identificaci[oó]n",
        r"Clase\s+de\s+documento"
    ],
}

# ---------------- Utilidades ----------------
EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.I)
PHONE_RE = re.compile(r"(\+?\d[\d\s\-()]{6,}\d)")

ADDRESS_HINTS = [
    "CALLE","CRA","CARRERA","AVENIDA","AV.","AV ","TRANSVERSAL","TV","DG","DIAGONAL","#"," N°"," NO."," NO ","-"
]
DUMMY_CONSONANT_RE = re.compile(r'^[A-Z]{3,}$')
VOWELS = set("AEIOUÁÉÍÓÚaeiouáéíóú")
PHONE_LABELS = {"Teléfono de notificación","Teléfono móvil","Teléfono Oficina"}

# Catálogo y normalizador de tipos de documento
DOC_TYPES_STD = {
    "CEDULA DE CIUDADANIA": "CÉDULA DE CIUDADANÍA",
    "CEDULA DE CIUDADANÍA": "CÉDULA DE CIUDADANÍA",
    "CÉDULA DE CIUDADANIA": "CÉDULA DE CIUDADANÍA",
    "CÉDULA DE CIUDADANÍA": "CÉDULA DE CIUDADANÍA",
    "CEDULA DE EXTRANJERIA": "CÉDULA DE EXTRANJERÍA",
    "CÉDULA DE EXTRANJERIA": "CÉDULA DE EXTRANJERÍA",
    "CÉDULA DE EXTRANJERÍA": "CÉDULA DE EXTRANJERÍA",
    "TARJETA DE IDENTIDAD": "TARJETA DE IDENTIDAD",
    "PASAPORTE": "PASAPORTE",
    "NIT": "NIT",
    "REGISTRO CIVIL": "REGISTRO CIVIL",
    "CARNET DIPLOMATICO": "CARNÉ DIPLOMÁTICO",
    "CARNÉ DIPLOMATICO": "CARNÉ DIPLOMÁTICO",
    "CARNET DIPLOMÁTICO": "CARNÉ DIPLOMÁTICO",
    "CARNÉ DIPLOMÁTICO": "CARNÉ DIPLOMÁTICO",
}

# Catálogo mínimo DANE (lo justo para este caso; ampliable)
DANE_MUNICIPIOS = {
    "11001": "BOGOTÁ, D.C.",
    "05001": "MEDELLÍN",
}


def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s or "") if unicodedata.category(c) != "Mn")

def _std(s: str) -> str:
    if not s: return s
    key = strip_accents(str(s)).upper().strip()
    key = re.sub(r"\s+", " ", key)
    return DOC_TYPES_STD.get(key, s.strip())

def cleanspace(s: str) -> str:
    if s is None: return ""
    s = s.replace("\xa0"," ").replace("\u200b"," ")
    s = re.sub(r"[\t\r\f\v]"," ", s)
    s = re.sub(r"\s+"," ", s)
    return s.strip()

# --------- Lugar: mejoras ----------

def is_address_like_line(line: str) -> bool:
    t = strip_accents((line or "").upper())
    return any(h in t for h in ADDRESS_HINTS)

def vowel_density_ratio(text: str) -> float:
    tokens = re.findall(r'[A-Za-zÁÉÍÓÚáéíóú]+', text or '')
    if not tokens: return 0.0
    has_vowel = sum(1 for t in tokens if any(ch in VOWELS for ch in t))
    return has_vowel / max(1, len(tokens))

def is_dummy_text(text: str) -> bool:
    if not text: return False
    vd = vowel_density_ratio(text)
    looks_addr = any(is_address_like_line(ln.strip()) for ln in (text or "").splitlines())
    allcaps_blocks = sum(1 for t in re.findall(r'[A-Z]{3,}', strip_accents(text)) if DUMMY_CONSONANT_RE.match(t))
    # Suavizamos el criterio: solo marcamos DUMMY si NO parece dirección y
    # además hay demasiados bloques ALL CAPS o la densidad vocálica es baja.
    if (not looks_addr) and (vd < 0.30 or allcaps_blocks >= 8):
        return True
    return False

def refine_lugar(value: str) -> str:
    if not value: return value
    raw_lines = [ln.strip() for ln in value.splitlines() if ln.strip()]
    addr_lines = [ln for ln in raw_lines if is_address_like_line(ln)]
    if not addr_lines: addr_lines = raw_lines[:2]
    lugar = " ".join(addr_lines[:2])
    if len(lugar) > 220:
        lugar = lugar[:220].rsplit(" ", 1)[0]
    return lugar.strip()

def resolve_lugar_fallback(raw: str):
    """Devuelve (texto, quality_tag) cuando el lugar fue marcado DUMMY.
    - Caso código DANE puro: 5 dígitos -> "05001 (MEDELLÍN)" si está en catálogo.
    - Caso cadena con 'BARRIO/LOCALIDAD/COMUNA' y 'BOGOTÁ, D.C.': sintetiza.
    - En caso contrario, devuelve el `raw` truncado legible con etiqueta HEURISTIC.
    """
    if not raw: return "", "EMPTY"
    raw_clean = cleanspace(raw)
    # DANE puro
    m = re.fullmatch(r"\s*(\d{5})\s*", raw_clean)
    if m:
        code = m.group(1)
        name = DANE_MUNICIPIOS.get(code)
        if name:
            return f"{code} ({name})", "CODE_ONLY"
        return code, "CODE_ONLY"

    # Patrón Bogotá con BARRIO/LOCALIDAD/COMUNA
    up = strip_accents(raw_clean).upper()
    if "BARRIO/LOCALIDAD/COMUNA" in up:
        barrio = None
        m1 = re.search(r"BARRIO/LOCALIDAD/COMUNA\s*:\s*([^/,:]+)", up)
        if m1: barrio = m1.group(1).title()
        loc = None
        m2 = re.search(r"LOCALIDAD\s+([^,/:]+)", up)
        if m2: loc = m2.group(1).title()
        bog = None
        m3 = re.search(r"BOGOTA\s*,?\s*D\.?C\.?", up)
        if m3: bog = "Bogotá, D.C."
        parts = [p for p in [barrio, loc, bog] if p]
        if parts:
            return ", ".join(parts), "HEURISTIC"
    # Fallback genérico – devolvemos algo legible, recortado
    text = raw_clean
    if len(text) > 140:
        text = text[:140].rsplit(' ', 1)[0]
    return text, "HEURISTIC"

# ---------------- Lectura PDF / OCR ----------------

def read_pdf_text(pdf_path: str) -> str:
    reader = PdfReader(pdf_path)
    chunks = []
    for page in reader.pages:
        try: t = page.extract_text() or ""
        except Exception: t = ""
        chunks.append(t)
    return "\n".join(chunks).strip()

def ocr_pdf(pdf_path: str, tmp_img_dir: str) -> str:
    if not OCR_OK:
        raise RuntimeError("OCR no disponible.")
    Path(tmp_img_dir).mkdir(parents=True, exist_ok=True)
    pages = convert_from_path(pdf_path, dpi=300, poppler_path=POPPLER_PATH)
    all_text = []
    for idx, img in enumerate(pages, start=1):
        img = img.convert("L"); img = ImageOps.autocontrast(img); img = img.filter(ImageFilter.SHARPEN)
        txt = pytesseract.image_to_string(img, lang="spa", config="--psm 6")
        if len(txt.strip()) < 10:
            txt = pytesseract.image_to_string(img, lang="eng", config="--psm 6")
        all_text.append(txt)
        try: img.save(os.path.join(tmp_img_dir, f"page_{idx:03d}.png"))
        except Exception: pass
    return "\n\n---- NUEVA PAGINA ----\n\n".join(all_text)

# ---------------- Split multi-caso (21 dígitos) ----------------
CASE_HEADER_LOOSE = re.compile(r"(?i)caso\s+noticia\s*[:：]\s*([^\n\r]{0,80})")
OCR_FIX_MAP = str.maketrans({'O':'0','o':'0','I':'1','l':'1','S':'5','B':'8'})

def sanitize_case_number(fragment: str) -> str:
    frag = fragment.replace('\xa0',' ').replace('\u200b',' ')
    frag = frag.translate(OCR_FIX_MAP)
    digits = re.sub(r"\D","", frag)
    return digits

def normalize_labels_multiline(text: str) -> str:
    t = text
    for label, patterns in LABEL_ALIASES.items():
        for pat in patterns:
            t = re.sub(rf"(?im){pat}\s*[:：]\s*", f"{label}: ", t)
    return t

def find_valid_case_headers(full_text: str):
    text = normalize_labels_multiline(full_text)
    accepted, rejected = [], []
    for m in CASE_HEADER_LOOSE.finditer(text):
        tail = m.group(1)
        num = sanitize_case_number(tail)
        if len(num) == 21:
            accepted.append((m.start(), num))
        else:
            rejected.append({"pos": m.start(), "snippet": (tail or '').strip(), "digits": num, "len": len(num)})
    accepted.sort(key=lambda x: x[0])
    return text, accepted, rejected


def split_cases(full_text: str) -> list:
    text, accepted, rejected = find_valid_case_headers(full_text)
    if not accepted:
        return [text]
    starts = [pos for pos,_ in accepted]
    blocks = []
    for i, s in enumerate(starts):
        e = starts[i+1] if i+1 < len(starts) else len(text)
        blocks.append(text[s:e].strip())
    return blocks


def extract_case_id(block: str):
    m = CASE_HEADER_LOOSE.search(block)
    if not m: return None
    num = sanitize_case_number(m.group(1))
    return num if len(num) == 21 else None

# ---------------- Extractores ----------------

def label_patterns(labels):
    pats = []
    for lb in labels:
        candidates = [rf"{re.escape(lb)}"] + LABEL_ALIASES.get(lb, [])
        for c in candidates:
            pats.append(re.compile(rf"(?is)\b{c}\s*[:：]"))
    return pats

ALL_LABELS = CASE_FIELDS + PERSON_FIELDS
ALL_LABEL_PATS = label_patterns(ALL_LABELS)
EXTRA_STOPS = ["Personas Vinculadas al Caso", "Información del Caso"]


def find_next_label_pos(text: str) -> int:
    next_pos = None
    for rp in (ALL_LABEL_PATS + label_patterns(EXTRA_STOPS)):
        m = rp.search(text)
        if m:
            p = m.start()
            if next_pos is None or p < next_pos:
                next_pos = p
    return next_pos if next_pos is not None else -1


def extract_between(text: str, start_label: str, stop_labels: list) -> str:
    starts = [re.search(rf"(?is)\b{pat}\s*[:：]", text) for pat in [re.escape(start_label)] + LABEL_ALIASES.get(start_label, [])]
    starts = [m for m in starts if m]
    if not starts:
        return ""
    s = min(starts, key=lambda m: m.start())
    sub = text[s.end():]
    stop_cands = []
    for sl in (stop_labels + EXTRA_STOPS):
        for pat in [re.escape(sl)] + LABEL_ALIASES.get(sl, []):
            stop_cands.append(re.compile(rf"(?is)\b{pat}\s*[:：]"))
    end_idx = len(sub)
    for rp in stop_cands:
        m2 = rp.search(sub)
        if m2 and m2.start() < end_idx:
            end_idx = m2.start()
    return cleanspace(sub[:end_idx])


def _find_label_positions(text: str, label: str):
    patterns = [re.compile(rf"(?is)\b{re.escape(label)}\s*[:：]")]
    for alias in LABEL_ALIASES.get(label, []):
        patterns.append(re.compile(rf"(?is)\b{alias}\s*[:：]"))
    matches = [p.search(text) for p in patterns]
    matches = [m for m in matches if m]
    return min(matches, key=lambda m: m.start()) if matches else None


def extract_between_strict(text: str, label: str, stop_labels: list) -> str:
    v = extract_between(text, label, stop_labels)
    if v:
        return v
    start_m = _find_label_positions(text, label)
    if not start_m:
        return ""
    sub = text[start_m.end():]
    stop_cands = []
    for sl in (stop_labels + EXTRA_STOPS):
        stop_cands.extend(
            [re.compile(rf"(?is)\b{re.escape(sl)}\s*[:：]")]
            + [re.compile(rf"(?is)\b{pat}\s*[:：]") for pat in LABEL_ALIASES.get(sl, [])]
        )
    end_idx = len(sub)
    for rp in stop_cands:
        m2 = rp.search(sub)
        if m2 and m2.start() < end_idx:
            end_idx = m2.start()
    return cleanspace(sub[:end_idx])

# ---------------- Emails / Phones ----------------

def extract_clean_email(raw: str) -> str:
    if not raw: return ""
    s = re.sub(r"\[[^\]]*\]\(mailto:([^\)]+)\)", r"\1", raw, flags=re.I)
    s = s.replace("mailto:", "")
    s_compact = s.replace(" ", "")
    m = EMAIL_RE.search(s_compact)
    if m:
        return m.group(0)
    m2 = EMAIL_RE.search(s)
    return m2.group(0) if m2 else ""
# ---------------- Parseo del caso y personas ----------------
STOP_AFTER_RELATO = [
    "Municipio Fiscal","Seccional","Unidad de Fiscalía","Despacho",
    "Estado de la asignación","Unidad de Enrutamiento","Estado del caso","Etapa del caso",
    "Calidad","Caso Noticia"
]


def parse_case_and_people(raw_text: str) -> dict:
    text = normalize_labels_multiline(raw_text.replace("\xa0"," ").replace("\u200b"," "))

    # ----- Campos del caso -----
    caso = {k: "" for k in CASE_FIELDS}
    for k in CASE_FIELDS:
        v = extract_between(text, k, ALL_LABELS)
        if not v:
            m = re.search(rf"(?im)\b{re.escape(k)}\s*[:：]\s*(.+)$", text)
            v = cleanspace(m.group(1)) if m else ""
        if k == "Etapa del caso":
            for extra in EXTRA_STOPS:
                v = re.split(rf"(?is)\b{re.escape(extra)}\s*[:：]", v)[0].strip()
        caso[k] = v

    # Lugar – con fallback cuando el detector lo marca DUMMY
    lugar_raw = caso.get("Lugar de los hechos", "")
    if lugar_raw:
        if is_dummy_text(lugar_raw):
            text2, tag = resolve_lugar_fallback(lugar_raw)
            caso["Raw_Lugar de los hechos"] = lugar_raw
            caso["Quality_Lugar de los hechos"] = tag
            caso["Lugar de los hechos"] = text2
        else:
            ref = refine_lugar(lugar_raw)
            caso["Raw_Lugar de los hechos"] = lugar_raw
            caso["Quality_Lugar de los hechos"] = "OK" if ref else "EMPTY"
            caso["Lugar de los hechos"] = ref
    else:
        caso["Raw_Lugar de los hechos"] = ""
        caso["Quality_Lugar de los hechos"] = "EMPTY"

    # Relato
    relato = extract_between(text, "Relato de los hechos", STOP_AFTER_RELATO)
    if not relato:
        m_q = re.search(r"(?s)¿.+$", text)
        if m_q:
            tail = m_q.group(0)
            pos = find_next_label_pos(tail)
            relato = tail[:pos] if pos != -1 else tail
    caso["Relato de los hechos"] = cleanspace(relato or caso.get("Relato de los hechos", ""))

    # SI/NO
    def _norm_si_no(raw):
        t = strip_accents(str(raw or "")).upper()
        if re.search(r"\bSI\b", t): return "SI"
        if re.search(r"\bNO\b", t): return "NO"
        if re.fullmatch(r"\s*S\s*", t): return "SI"
        if re.fullmatch(r"\s*N\s*", t): return "NO"
        return ""
    caso["Procedimiento Abreviado?"] = _norm_si_no(caso.get("Procedimiento Abreviado?", ""))
    caso["Priorizado"] = _norm_si_no(caso.get("Priorizado", ""))

    # ----- Personas -----
    personas = []
    chunks = re.split(r"(?im)^\s*Calidad\s*[:：]", text)
    for ch in chunks[1:]:
        pdata = {lbl: "" for lbl in PERSON_FIELDS}
        first_line = ch.splitlines()[0] if ch.strip() else ""
        pdata["Calidad"] = cleanspace(first_line)

        # Buscaremos la posición de cada etiqueta para poder acotar vecindad
        all_stop_labels = PERSON_FIELDS

        for lbl in PERSON_FIELDS[1:]:
            val = extract_between_strict(ch, lbl, PERSON_FIELDS)

            # --- Teléfonos: vecindad acotada entre etiqueta actual y la siguiente ---
            if (not val) and (lbl in PHONE_LABELS):
                # pos de etiqueta actual
                start_m = _find_label_positions(ch, lbl)
                start = start_m.end() if start_m else 0
                # pos de la siguiente etiqueta (mínima mayor a start)
                next_pos = None
                for lab in PERSON_FIELDS:
                    m2 = _find_label_positions(ch, lab)
                    if m2 and m2.start() > start:
                        if next_pos is None or m2.start() < next_pos:
                            next_pos = m2.start()
                end = next_pos if next_pos is not None else min(len(ch), start + 180)
                vicinity = ch[start:end]
                mphone = PHONE_RE.search(vicinity)
                if mphone:
                    val = cleanspace(mphone.group(1))

            # Correos
            if lbl == "Correo Electrónico":
                email = extract_clean_email(val)
                if not email:
                    start_m = _find_label_positions(ch, lbl)
                    start = start_m.end() if start_m else 0
                    next_pos = None
                    for lab in PERSON_FIELDS:
                        m2 = _find_label_positions(ch, lab)
                        if m2 and m2.start() > start:
                            if next_pos is None or m2.start() < next_pos:
                                next_pos = m2.start()
                    end = next_pos if next_pos is not None else min(len(ch), start + 220)
                    vicinity = ch[start:end]
                    email = extract_clean_email(vicinity)
                val = email

            # Heurística/normalización para tipo de documento
            if lbl == "Documento":
                # Si no vino, lo inferimos del bloque antes de "Número documento"
                if not val:
                    m_num = _find_label_positions(ch, "Número documento")
                    region = ch[:m_num.start()] if m_num else ch
                    for cand in DOC_TYPES_STD.keys():
                        rx = re.compile(re.escape(cand), re.I)
                        mm = rx.search(strip_accents(region))
                        if mm:
                            val = _std(mm.group(0)); break
                else:
                    val = _std(val)

            # Evitar que un teléfono duplique documento
            if lbl in PHONE_LABELS and val:
                only_digits = re.sub(r"\D", "", val)
                doc_vals = [
                    re.sub(r"\D","", str(pdata.get("Documento",""))),
                    re.sub(r"\D","", str(pdata.get("Número documento","")))
                ]
                if only_digits and only_digits in doc_vals:
                    val = ""

            pdata[lbl] = val

        # Regla post: si Teléfono de notificación == Teléfono móvil => vaciamos fijo
        try:
            if pdata.get("Teléfono de notificación") and pdata.get("Teléfono móvil") and \
               re.sub(r"\D","", str(pdata.get("Teléfono de notificación"))) == re.sub(r"\D","", str(pdata.get("Teléfono móvil"))):
                pdata["Teléfono de notificación"] = ""
        except Exception:
            pass

        if any(str(v).strip().lower() not in ("", "nan", "-") for v in pdata.values()):
            personas.append(pdata)

    fila = dict(caso)
    for i, p in enumerate(personas, start=1):
        suf = f"_{i}"
        for lbl in PERSON_FIELDS:
            fila[f"{lbl}{suf}"] = p.get(lbl, "")
    return fila

# ---------------- Reorden por rol (INDICIADO primero) ----------------
ROLE_PRIORITY_REST = ["DENUNCIANTE", "VICTIMA", "TESTIGO"]

def normalize_role(val):
    try:
        import pandas as _pd
        if _pd.isna(val): s = ""
        else: s = str(val)
    except Exception:
        s = str(val) if val is not None else ""
    v = strip_accents(s).upper().strip()
    if v.startswith("INDICIADO"): return "INDICIADO"
    if v.startswith("DENUNCIANTE"): return "DENUNCIANTE"
    if v.startswith("VICTIMA"): return "VICTIMA"
    if v.startswith("TESTIGO"): return "TESTIGO"
    return v


def collect_person_blocks_from_row(row: pd.Series) -> list:
    ns = sorted({int(m.group(1)) for c in row.index for m in [re.search(r"_(\d+)$", str(c))] if m})
    def _nonempty(x):
        try:
            import pandas as _pd
            if _pd.isna(x): return False
        except Exception: pass
        s = str(x).strip(); return s not in ("", "nan", "-")
    persons = []
    for n in ns:
        block = {f: row.get(f"{f}_{n}", "") for f in PERSON_FIELDS}
        if any(_nonempty(v) for v in block.values()):
            block["__role__"] = normalize_role(block.get("Calidad", ""))
            persons.append(block)
    return persons


def write_person_blocks_to_row(base: dict, persons_by_new_index: list):
    out = dict(base)
    for new_idx, p in persons_by_new_index:
        for f in PERSON_FIELDS:
            out[f"{f}_{new_idx}"] = p.get(f, "")
    return out


def reorder_and_expand(df: pd.DataFrame) -> pd.DataFrame:
    per_row = []
    for _, r in df.iterrows():
        base = {k: r.get(k, "") for k in r.index if not re.search(r"_(\d+)$", str(k))}
        persons = collect_person_blocks_from_row(r)
        indic = [p for p in persons if p.get("__role__") == "INDICIADO"]
        rest = [p for p in persons if p.get("__role__") != "INDICIADO"]
        def rest_key(p):
            rp = p.get("__role__", "")
            try: return (ROLE_PRIORITY_REST.index(rp), rp)
            except ValueError: return (len(ROLE_PRIORITY_REST)+1, rp)
        rest_sorted = sorted(rest, key=rest_key)
        per_row.append((base, indic, rest_sorted))

    max_n = 0
    k_max = 0
    for base, indic, rest_sorted in per_row:
        m = len(indic) + len(rest_sorted)
        max_n = max(max_n, m)
        k_max = max(k_max, len(indic))
    if max_n == 0: max_n = 1

    new_rows = []
    for base, indic, rest_sorted in per_row:
        pairs = []
        for i, p in enumerate(indic, start=1): pairs.append((i, p))
        start_idx = k_max + 1
        for j, p in enumerate(rest_sorted, start=start_idx): pairs.append((j, p))
        new_rows.append(write_person_blocks_to_row(base, pairs))

    df2 = pd.DataFrame(new_rows)

    def person_cols(i):
        return [
            f"Calidad_{i}", f"Documento_{i}", f"Número documento_{i}", f"Nombre_{i}",
            f"Departamento de notificación_{i}", f"Municipio de notificación_{i}",
            f"Dirección de notificación_{i}", f"Teléfono de notificación_{i}",
            f"Teléfono móvil_{i}", f"Correo Electrónico_{i}", f"Teléfono Oficina_{i}"
        ]

    base_cols = ["Fecha de los Hechos","Caso Noticia","Seccional","Unidad de Fiscalía","Despacho","Unidad de Enrutamiento","Delito"]
    pre_relato = ["Caracterización","Modalidad","Modo","Municipio Fiscal","Lugar de los hechos"]
    relato = ["Relato de los hechos"]
    post_relato_pre_grado = ["Estado del caso","Etapa del caso","Estado de la asignación","Ley de Aplicabilidad","Procedimiento Abreviado?","Priorizado","Tipo Noticia"]
    grado = ["Grado Delito"]

    desired = []
    desired += base_cols + pre_relato + relato
    for i in range(1, k_max+1): desired += person_cols(i)
    desired += post_relato_pre_grado + grado
    for i in range(k_max+1, max_n+1): desired += person_cols(i)

    for col in desired:
        if col not in df2.columns: df2[col] = ""
    restantes = [c for c in df2.columns if c not in desired]
    final_cols = desired + restantes
    return df2.reindex(columns=final_cols)

# ---------------- Post-proceso ----------------
DATE_COLS = {"Fecha de los Hechos"}
NAME_COL_PREFIX = "Nombre_"
NUMDOC_PREFIX = "Número documento_"  # solo estos a dígitos

# --- FIX FECHA-HORA (inserta espacio si falta entre fecha y hora) ---
import re as _re

def _ensure_space_datetime(s: str):
    if not s: return s
    s2 = str(s)
    # dd/mm/yyyyHH:MM:SS -> dd/mm/yyyy HH:MM:SS
    s2 = _re.sub(r"^(\d{2}/\d{2}/\d{4})(\d{2}:\d{2}:\d{2})$", r"\1 \2", s2)
    # dd-mm-yyyyHH:MM:SS -> dd-mm-yyyy HH:MM:SS
    s2 = _re.sub(r"^(\d{2}-\d{2}-\d{4})(\d{2}:\d{2}:\d{2})$", r"\1 \2", s2)
    return s2



def _remove_spaces_between_digits(text: str) -> str:
    return re.sub(r"(?<=\d)\s+(?=\d)", "", str(text))

def _fix_datetime_cell(s: str):
    if not s: return s
    # Garantizamos el espacio fecha-hora si viene pegado (p.ej., 07/06/202400:36:00)
    s2 = _ensure_space_datetime(str(s))
    # Compactamos espacios extra entre dígitos sin afectar el espacio fecha-hora
    s2 = _remove_spaces_between_digits(s2)
    s2 = _ensure_space_datetime(s2)
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y %H:%M:%S", "%d-%m-%Y"):
        try:
            dt = datetime.datetime.strptime(s2, fmt)
            return dt.strftime("%d/%m/%Y %H:%M:%S")
        except Exception:
            continue
    return s2


def _digits_only(s: str) -> str:
    return re.sub(r"\D", "", str(s))

def _normalize_phone(s: str) -> str:
    if not s: return s
    s1 = str(s).strip()
    plus = s1.startswith("+")
    digits = _digits_only(s1)
    return ("+"+digits) if plus else digits

def _fix_broken_upper_name(s: str) -> str:
    if not s: return s
    raw = str(s)
    if re.search(r"[a-zñáéíóú]", raw):
        return raw.strip()
    toks = raw.strip().split()
    if not toks: return raw.strip()
    out = []
    for tok in toks:
        if out and re.fullmatch(r"[A-ZÁÉÍÓÚÑ]{1,2}", tok):
            out[-1] = (out[-1] + tok)
        else:
            out.append(tok)
    return " ".join(out)

INTRUSIVE_LABELS_IN_NAME = re.compile(
    r"(?:\bDepartamento\s+de\s+notificaci[oó]n\s*:\s*.*)|(?:\bMunicipio\s+de\s+notificaci[oó]n\s*:\s*.*)",
    re.IGNORECASE | re.DOTALL
)

def _strip_injected_labels_from_name(s: str) -> str:
    if not s: return s
    s2 = INTRUSIVE_LABELS_IN_NAME.split(str(s))[0]
    return s2.strip()


def postprocess_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df2 = df.copy()
    # Fechas
    for col in df2.columns:
        if col in DATE_COLS:
            df2[col] = df2[col].apply(_fix_datetime_cell)
    # Compactar números embebidos
    def compact_numeric_runs(x):
        if not isinstance(x, str):
            try:
                return x if not pd.isna(x) else ""
            except Exception:
                return x
        return re.sub(r"(?<=\d)\s+(?=\d)", "", x)
    objcols = df2.select_dtypes(include=["object", "string"]).columns
    # 👇 Línea NUEVA: no compactar en columnas de fecha
    objcols = [c for c in objcols if c not in DATE_COLS]
    
    for col in objcols:
        df2[col] = df2[col].apply(compact_numeric_runs)
    # Teléfonos
    for col in df2.columns:
        base = col.split("_")[0]
        if base in PHONE_LABELS:
            df2[col] = df2[col].apply(_normalize_phone)
    # Documentos
    for col in df2.columns:
        if col.startswith(NUMDOC_PREFIX):
            df2[col] = df2[col].apply(_digits_only)
        elif col.startswith("Documento_"):
            df2[col] = df2[col].apply(_std)
    # Nombres
    for col in df2.columns:
        if col.startswith(NAME_COL_PREFIX):
            df2[col] = df2[col].apply(_fix_broken_upper_name)
            df2[col] = df2[col].apply(_strip_injected_labels_from_name)
    
    # (Opcional pero recomendado) Reasegurar formato de fecha-hora
    for col in df2.columns:
         if col in DATE_COLS:
            df2[col] = df2[col].apply(_fix_datetime_cell)

    return df2


def safe_excel_path(base_path: str) -> str:
    try:
        f = open(base_path, 'ab'); f.close()
        return base_path
    except PermissionError:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        root, ext = os.path.splitext(base_path)
        alt = f"{root} ({ts}){ext}"
        print(f"Archivo de salida bloqueado, guardando como: {alt}")
        return alt


def write_excel_split(df: pd.DataFrame, out_path: str):
    try:
        obj_cols = df.select_dtypes(include=["object", "string"]).columns.tolist()
        for c in obj_cols:
            df[c] = df[c].apply(lambda x: x.strip() if isinstance(x, str) else x)
    except Exception:
        pass

    raw_quality_cols = [c for c in df.columns if c.startswith("Raw_") or c.startswith("Quality_")]
    meta_cols = [c for c in ["CaseId21_Valid","Confidence_PageMean","Extraction_Method"] if c in df.columns]
    datos_cols = [c for c in df.columns if c not in raw_quality_cols + meta_cols]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df[datos_cols].to_excel(writer, sheet_name="Datos", index=False)
        ws = writer.sheets["Datos"]
        for ci, col in enumerate(df[datos_cols].columns, start=1):
            values = [str(col)] + [str(v) for v in df[datos_cols][col].tolist()]
            maxlen = max(len(v) for v in values)
            ws.column_dimensions[get_column_letter(ci)].width = min(max(12, int(maxlen*0.9)), 85)
        if "Relato de los hechos" in df[datos_cols].columns:
            cidx = df[datos_cols].columns.get_loc("Relato de los hechos") + 1
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=cidx, max_col=cidx):
                for cell in row: cell.alignment = Alignment(wrap_text=True, vertical="top")
        for cell in ws[1]: cell.font = Font(bold=True)
        ws.freeze_panes = "A2"

        if raw_quality_cols or meta_cols:
            df[raw_quality_cols + meta_cols].to_excel(writer, sheet_name="Auditoría", index=False)
            ws2 = writer.sheets["Auditoría"]
            for ci, col in enumerate((raw_quality_cols + meta_cols), start=1):
                values = [str(col)] + [str(v) for v in df[col].tolist()]
                maxlen = max(len(v) for v in values)
                ws2.column_dimensions[get_column_letter(ci)].width = min(max(12, int(maxlen*0.9)), 85)
            for cell in ws2[1]: cell.font = Font(bold=True)
            ws2.freeze_panes = "A2"

# ---------------- MAIN ----------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-i','--input', default=DEFAULT_PDF_NAME, help='Nombre del PDF de entrada')
    parser.add_argument('-o','--output', default=DEFAULT_XLSX_NAME, help='Nombre del Excel de salida')
    args = parser.parse_args()

    pdf_path = args.input if os.path.isabs(args.input) else os.path.join(BASE_DIR, args.input)
    xlsx_path = args.output if os.path.isabs(args.output) else os.path.join(BASE_DIR, args.output)
    log_txt = os.path.join(BASE_DIR, "ocr_text.txt")
    tmp_img = os.path.join(BASE_DIR, "_tmp_pdf_imgs")

    if not os.path.isfile(pdf_path):
        raise FileNotFoundError(f"No se encontró el PDF: {pdf_path}")

    text_all = read_pdf_text(pdf_path)
    used_ocr = False
    if len(text_all.strip()) < 20:
        used_ocr = True
        text_all = ocr_pdf(pdf_path, tmp_img)

    try:
        with open(log_txt, "w", encoding="utf-8") as f: f.write(text_all)
    except Exception: pass

    case_chunks = split_cases(text_all)

    if DEDUPE_BY_ID:
        seen = set(); chunks_use = []
        for ch in case_chunks:
            cid = extract_case_id(ch)
            if not cid or cid in seen: continue
            seen.add(cid); chunks_use.append(ch)
    else:
        chunks_use = case_chunks

    filas = []
    for idx, chunk in enumerate(chunks_use, start=1):
        fila = parse_case_and_people(chunk)
        filas.append(fila)

    df = pd.DataFrame(filas)
    df = reorder_and_expand(df)
    df = postprocess_dataframe(df)

    out_path = safe_excel_path(xlsx_path)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    write_excel_split(df, out_path)
    try:
        if hasattr(os, "startfile"):
            os.startfile(out_path)
    except Exception:
        pass
    print(f"Archivo generado: {out_path} (modo: {'OCR' if used_ocr else 'Texto directo'})")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("\n*** ERROR EJECUTANDO ESTRUCTURADOR (FIX3+7) ***")
        print(str(e))
        traceback.print_exc()
        sys.exit(1)